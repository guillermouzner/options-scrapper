import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl


# carga el archivo Excel existente
workbook = openpyxl.load_workbook('D:\pruebaBonos.xlsx')

# selecciona la hoja donde se escribirán los datos
sheet = workbook['Hoja1']

# especifica la celda de inicio para escribir los datos
start_row = 500
start_column = 1

# borra las filas existentes
sheet.delete_rows(500, sheet.max_row)

# Inicializar el controlador del navegador
driver = webdriver.Chrome()

# Navegar a la página web
driver.get("https://www.bcra.gob.ar/")

# Encontrar el elemento <ul> con la clase "nav navbar-nav"
nav_menu = driver.find_element(By.CLASS_NAME, "nav.navbar-nav")

# Encontrar todos los elementos <a> dentro del elemento <ul>
links = nav_menu.find_elements(By.TAG_NAME, "a")

# Buscar el enlace "Estadísticas" dentro de la lista de enlaces y hacer clic en él
for link in links:
    if link.text == "Estadísticas":
        link.click()
        break

# Esperar a que aparezca el menú desplegable "Estadísticas"
wait = WebDriverWait(driver, 10)
estadisticas_menu = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "ul.dropdown-menu.dropdown-menu-large.row")))

# Encontrar el elemento <a> "Principales variables" dentro del menú desplegable "Estadísticas"
principales_variables_link = estadisticas_menu.find_element(By.XPATH, "//li[@class='dropdown-header']/a[text()='Principales variables']")

# Hacer clic en el enlace "Principales variables"
ActionChains(driver).move_to_element(principales_variables_link).click().perform()

# Encontrar la tabla de estadísticas
estadisticas_table = driver.find_element(By.CLASS_NAME, "table-BCRA")

# Encontrar el enlace de la Tasa de Política Monetaria dentro de la tabla y hacer clic en él
link = estadisticas_table.find_element(By.XPATH, "//a[contains(@href,'serie=1222&detalle=BADLAR en pesos de bancos privados (en % n.a.)')]")
link.click()

# Encontrar los elementos input de fecha_desde y fecha_hasta y escribir las fechas
fecha_desde_input = driver.find_element(By.NAME, "fecha_desde")
fecha_desde_input.clear() # Limpiar cualquier valor anterior
fecha_desde_input.send_keys("12/01/2023")

fecha_hasta_input = driver.find_element(By.NAME, "fecha_hasta")
fecha_hasta_input.clear() # Limpiar cualquier valor anterior
fecha_hasta_input.send_keys("24/02/2023")

# Encontrar el botón "Consultar" y hacer clic en él
consultar_button = driver.find_element(By.NAME, "B1")
consultar_button.click()

# Encontrar la tabla de estadísticas
estadisticas_table = driver.find_element(By.CLASS_NAME, "table-BCRA")

# Obtener el HTML de la tabla
tabla_html = estadisticas_table.get_attribute('outerHTML')

# Convertir la tabla HTML en un dataframe de pandas
df = pd.read_html(tabla_html)[0]

# Limpiar los valores numéricos (reemplazar puntos por comas y convertir a float)
df['Valor'] = df['Valor'].replace({',': '.', '\.': ''}, regex=False).astype(float) / 10000



# convierte el dataframe en una lista de listas
data = df.values.tolist()

# escribe los datos en el archivo Excel
for row in range(len(data)):
    for column in range(len(data[row])):
        sheet.cell(row=start_row+row, column=start_column+column, value=data[row][column])

# guarda los cambios en el archivo Excel
workbook.save('D:\pruebaBonos.xlsx')

# Imprimir el dataframe
print("HECHO")

# Esperar 1 segundo
time.sleep(1)

# Cerrar el navegador
driver.quit()
